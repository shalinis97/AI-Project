import os
import PyPDF2
import pdfplumber
from docx import Document
import openpyxl
import pytesseract
from PIL import Image
#from tika import parser
from pdf2image import convert_from_path
import fitz  # PyMuPDF
import tempfile

from dotenv import load_dotenv
import os
import sqlite3


load_dotenv(override=True)  # Loads variables from .env into environment

reportDir = rf"{os.getenv('REPROT_DIR')}"

pytesseract.pytesseract.tesseract_cmd = rf"{os.getenv('WIN_TESSERACT_DIR')}"

#print("Report Directory:", reportDir)


def read_pdf(u_file):
     # Save to temp file for libraries that require a path
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as file:
        file.write(u_file.read())
        file.flush()
        file = file.name

    with open(file, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        pdf2_text = ""
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pdf2_text += text + "\n"
    #print("PyPDF2 Extracted Text:\n", pdf2_text)

            # pdfplumber extraction
    with pdfplumber.open(file) as pdf:
        pdfplumber_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pdfplumber_text += text + "\n"
    #print("pdfplumber Extracted Text:\n", pdfplumber_text)
    
   
    pdf_document = fitz.open(file)
    pytesseract_text = ""
    for page_num in range(len(pdf_document)):
        page = pdf_document[page_num]

        # Render page to image (300 DPI)
        pix = page.get_pixmap(dpi=300)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

        # OCR using pytesseract
        text = pytesseract.image_to_string(img)

        if text:
            pytesseract_text += text + "\n"

    print("pytesseract Extracted Text:\n", pytesseract_text)
    
    print("pdf Extracted Text pdf2_text len :", len(pdf2_text))
    print("pdf Extracted Text pdfplumber_text len :", len(pdfplumber_text))
    print("pdf Extracted Text pytesseract_text len :", len(pytesseract_text))


    text_data = max(pdf2_text, pdfplumber_text, pytesseract_text, key=len)

    print("-------------------------------------------------PDF----------------------------------------------------------------------------\n")

    return text_data



def read_doc(file):
    doc = Document(file)
    docx_text = "\n".join(para.text for para in doc.paragraphs)
    #print("DOCX Extracted Text:\n", docx_text)
    print("------------------------------------------------DOCX-----------------------------------------------------------------------------\n")

    return docx_text



def read_excel(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    excel_data = {}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        excel_data[sheetname] = [row for row in ws.iter_rows(values_only=True)]
    #print("Excel Data:\n", excel_data)
    print("---------------------------------------------EXCEL--------------------------------------------------------------------------------\n")

    flattened_text = ""
    for sheetname, rows in excel_data.items():
        flattened_text += f"Sheet: {sheetname}\n"
        flattened_text += "\n".join(
            " | ".join(str(cell).strip() for cell in row if cell is not None and str(cell).strip() != "")
            for row in rows
            if any(cell is not None and str(cell).strip() != "" for cell in row)
        )
        flattened_text += "\n\n"

    return flattened_text



def read_image(file):
    img = Image.open(file)
    ocr_text = pytesseract.image_to_string(img)
    #print("OCR Extracted Text:\n", ocr_text)
    print("---------------------------------------------OCR--------------------------------------------------------------------------------\n")

    return ocr_text



#apache tika
def read_file(file):
    # Parse the file
    parsed = parser.from_file(file)

    # Extract text content
    text = parsed['content']

    # Extract metadata
    metadata = parsed['metadata']

    print("Text Content:\n", text)
    print("Metadata:\n", metadata)

    return text,metadata

# def read_files(reportDir):
#     text_data = {}

#     for filename in os.listdir(reportDir):
#         if filename.startswith('~') or filename.startswith('.'):
#             continue  # Skip temp/hidden files

#         filepath = os.path.join(reportDir, filename)
#         print(f"\nProcessing: {filename}")
def read_files(uploaded_files):
    text_data = {}

    for uploaded_file in uploaded_files:
        filename = uploaded_file.name
        print(f"\nProcessing: {filename}")
        try:

            #text_data,metadata = read_file(filepath)

            # if(len(text_data)!=0): 
            #     return text_data

            if filename.lower().endswith(".docx"):
                text_data[filename.lower()] = read_doc(uploaded_file)

            elif filename.lower().endswith(".pdf"):
                text_data[filename.lower()] = read_pdf(uploaded_file)

            elif filename.lower().endswith(".xlsx"):
                text_data[filename.lower()] = read_excel(uploaded_file)

            elif filename.lower().endswith((".png", ".jpg", ".jpeg")):
                text_data[filename.lower()] = read_image(uploaded_file)

            else:
                print("Unsupported file type.")


        except Exception as e:
            print(f"Error processing {filename}: {e}")
    
    return text_data 




# --- DB functions ---
def init_db(db_path="file_data.db"):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT UNIQUE,
            filetype TEXT,
            content TEXT
        )
    ''')
    conn.commit()
    conn.close()

def get_file_from_db(filename, db_path="file_data.db"):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute('SELECT content FROM files WHERE filename=?', (filename,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else None

def save_to_db(filename, filetype, content, db_path="file_data.db"):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute('''
        INSERT OR REPLACE INTO files (filename, filetype, content)
        VALUES (?, ?, ?)
    ''', (filename, filetype, content))
    conn.commit()
    conn.close()
